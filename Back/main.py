from fastapi import FastAPI, HTTPException, Depends, Header, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, date, timedelta
from uuid import uuid4
import unicodedata
from pathlib import Path
import openpyxl
from pathlib import Path
from typing import Optional, List, Dict
from datetime import datetime, date, timedelta
import os
import json
from io import BytesIO

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

# ClÃ© dâ€™accÃ¨s admin (simple pour ton prototype)
ADMIN_TOKEN = "admin123"   # tu peux changer la valeur

def require_admin(Authorization: str = Header(None)):
    """
    VÃ©rifie que la requÃªte possÃ¨de le bon token admin.
    Exemple cÃ´tÃ© front : headers: { Authorization: 'Bearer admin123' }
    """
    if not Authorization:
        raise HTTPException(status_code=401, detail="Token manquant")
    token = Authorization.replace("Bearer ", "")
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=403, detail="AccÃ¨s refusÃ©")
    return True


app = FastAPI()

origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "https://projet-fjord-king.vercel.app",
    "https://projet-fjord-king-git-main-marrache-raphaels-projects.vercel.app",
    "https://projet-fjord-king-cmvq0fx63-marrache-raphaels-projects.vercel.app",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "employes_exemple.xlsx"
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

GOOGLE_DRIVE_FILE_ID = os.getenv("GOOGLE_DRIVE_FILE_ID")
GOOGLE_SERVICE_ACCOUNT_JSON = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
GOOGLE_SERVICE_ACCOUNT_JSON_PATH = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON_PATH")

def _get_drive_service():
    if not GOOGLE_DRIVE_FILE_ID:
        return None
    creds = None
    if GOOGLE_SERVICE_ACCOUNT_JSON:
        info = json.loads(GOOGLE_SERVICE_ACCOUNT_JSON)
        creds = service_account.Credentials.from_service_account_info(
            info, scopes=["https://www.googleapis.com/auth/drive"]
        )
    elif GOOGLE_SERVICE_ACCOUNT_JSON_PATH:
        creds = service_account.Credentials.from_service_account_file(
            GOOGLE_SERVICE_ACCOUNT_JSON_PATH, scopes=["https://www.googleapis.com/auth/drive"]
        )
    if not creds:
        return None
    return build("drive", "v3", credentials=creds)

def _download_excel_from_drive():
    if not GOOGLE_DRIVE_FILE_ID:
        return
    service = _get_drive_service()
    if not service:
        return
    try:
        request = service.files().get_media(fileId=GOOGLE_DRIVE_FILE_ID)
        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        EXCEL_FILE.write_bytes(fh.getvalue())
    except Exception as exc:
        print(f"Drive download failed: {exc}")

def _upload_excel_to_drive():
    if not GOOGLE_DRIVE_FILE_ID or not EXCEL_FILE.exists():
        return
    service = _get_drive_service()
    if not service:
        return
    try:
        media = MediaFileUpload(
            str(EXCEL_FILE),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        service.files().update(fileId=GOOGLE_DRIVE_FILE_ID, media_body=media).execute()
    except Exception as exc:
        print(f"Drive upload failed: {exc}")

_download_excel_from_drive()
# Liste en mÃ©moire pour stocker les demandes
REQUESTS = []
REQUEST_COUNTER = 1  # pour gÃ©nÃ©rer un ID unique Ã  chaque demande

_employees_cache: List[Dict] = []
_employees_mtime: float | None = None

def _read_employees_from_excel() -> List[Dict]:
    if not EXCEL_FILE.exists():
        _download_excel_from_drive()
    if not EXCEL_FILE.exists():
        raise RuntimeError(f"Fichier Excel introuvable: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active  # 1Ã¨re feuille

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    # On attend: Nom, Prenom, Email, Secteur, Code, CongesCumules, (Telephone optionnel)
    col = {h: i for i, h in enumerate(headers)}

    employees: List[Dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or (row[col["Nom"]] is None and row[col["Prenom"]] is None):
            continue
        nom = str(row[col["Nom"]] or "").strip()
        prenom = str(row[col["Prenom"]] or "").strip()
        email = str(row[col["Email"]] or "").strip()
        secteur = str(row[col["Secteur"]] or "").strip()
        code = str(row[col["Code"]] or "").strip()
        cumules = int(row[col.get("CongesCumules", -1)] or 0) if "CongesCumules" in col else 0
        solde = int(row[col.get("Solde", -1)] or 0) if "Solde" in col else 0
        tel = str(row[col["Telephone"]]).strip() if "Telephone" in col and row[col["Telephone"]] else None

        employees.append({
            "full_name": f"{prenom} {nom}",   # âš ï¸ ICI: identifiant saisi "Prenom Nom"
            "email": email,
            "sector": secteur,
            "code": code,
            "cumules": cumules,
            "solde": solde,
            "telephone": tel,
        })
    return employees

def _ensure_employees_loaded():
    global _employees_cache, _employees_mtime
    _download_excel_from_drive()
    mtime = EXCEL_FILE.stat().st_mtime if EXCEL_FILE.exists() else None
    if _employees_mtime != mtime:
        _employees_cache = _read_employees_from_excel()
        _employees_mtime = mtime

def get_employees() -> List[Dict]:
    _ensure_employees_loaded()
    return _employees_cache

# ================== "BASE" EMPLOYÃ‰S (simule ton Excel) ==================

BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "employes_exemple.xlsx"  # le fichier que tu as placÃ© Ã  cÃ´tÃ© de main.py


def load_employees_from_excel() -> list[dict]:
    if not EXCEL_FILE.exists():
        raise RuntimeError(f"Fichier Excel des employÃ©s introuvable : {EXCEL_FILE}")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active  # premiÃ¨re feuille

    employees: list[dict] = []

    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            # On saute la ligne d'en-tÃªte : Nom, Prenom, Email, Secteur, Code
            first = False
            continue

        nom, prenom, email, secteur, code, CongesCumules, Telephone,  = row

        # sÃ©curitÃ© au cas oÃ¹ il y a des lignes vides
        if not nom or not prenom:
            continue

        employees.append(
            {
                "full_name": f"{nom} {prenom}",
                "email": str(email).strip() if email else "",
                "sector": str(secteur).strip() if secteur is not None else "",
                "code": str(code).strip() if code is not None else "",
            }
        )

    return employees

employees = get_employees()  # relit si besoin via cache (inclut secteur + cumules)

def set_employee_cumules_in_excel(full_name: str, new_value: int) -> None:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    if "CongesCumules" not in col:
        raise RuntimeError("La colonne 'CongesCumules' est absente de l'Excel.")

    # Retrouver la ligne par "Prenom Nom"
    for r in range(2, ws.max_row + 1):
        nom = str(ws.cell(row=r, column=col["Nom"] + 1).value or "").strip()
        prenom = str(ws.cell(row=r, column=col["Prenom"] + 1).value or "").strip()
        if f"{prenom} {nom}".strip().lower() == full_name.strip().lower():
            ws.cell(row=r, column=col["CongesCumules"] + 1).value = int(new_value)
            wb.save(EXCEL_FILE)
            _upload_excel_to_drive()
            return
    raise RuntimeError(f"EmployÃ© '{full_name}' introuvable pour mise Ã  jour.")


# ================== MODELES ==================

class WorkerLogin(BaseModel):
    identifier: str  # "Nom Prenom" saisi par le travailleur
    code: str

class RequestCreate(BaseModel):
    worker_name: str
    worker_email: str
    sector: Optional[str] = None  # Secteur 1..5
    category: str                 # "vacation" ou "sick_leave"
    start_date: date
    end_date: date
    reason: Optional[str] = None
    # champs "anciens" qu'on garde pour compatibilitÃ© / affichage
    title: Optional[str] = None
    description: Optional[str] = None
    urgency: Optional[str] = "normal"

class RequestOut(RequestCreate):
    id: int
    status: str
    manager_comment: Optional[str] = None
    created_at: datetime

class RequestUpdate(BaseModel):
    status: Optional[str] = None
    manager_comment: Optional[str] = None
    
class RequestUpdate(BaseModel):
    status: Optional[str] = None
    manager_comment: Optional[str] = None

class LoginData(BaseModel):
    username: str
    password: str

class CumuleUpdate(BaseModel):
    days: int

# ================== STOCKAGE DEMANDES ==================

requests_db: List[RequestOut] = []
current_id = 1

MANAGERS = {
    "admin": "admin123"
}

FAKE_TOKEN = "super-secret-token"

def get_manager_token(authorization: str = Header(None)):
    if not authorization:
        raise HTTPException(status_code=401, detail="Token manquant")
    try:
        scheme, token = authorization.split()
    except ValueError:
        raise HTTPException(status_code=401, detail="Format de token invalide")

    if scheme.lower() != "bearer" or token != FAKE_TOKEN:
        raise HTTPException(status_code=401, detail="Token invalide")
    return True

def _days_inclusive(d1: date, d2: date) -> int:
    total = 0
    cur = d1
    while cur <= d2:
        if cur.weekday() != 6:  # 6 = Sunday
            total += 1
        cur = cur + timedelta(days=1)
    return total

def _to_date(value) -> date | None:
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None

def _norm_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    # Remove accents to avoid encoding mismatches (terminÃ©e vs terminÃ‡Â¸e).
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if ord(ch) < 128)

def _dates_in_range(start: date, end: date):
    cur = start
    while cur <= end:
        yield cur
        cur = cur + timedelta(days=1)

def _build_sector_day_counts(exclude_req=None):
    counts: Dict[tuple, int] = {}
    for r in REQUESTS:
        if exclude_req is not None and r is exclude_req:
            continue
        if _norm_text(r.get("status")) != "terminee":
            continue
        sector = r.get("sector") or ""
        if not sector:
            continue
        start = _to_date(r.get("start_date"))
        end = _to_date(r.get("end_date"))
        if not start or not end:
            continue
        for d in _dates_in_range(start, end):
            key = (sector, d.isoformat())
            counts[key] = counts.get(key, 0) + 1
    return counts

def _recompute_cumules_for(full_name: str) -> int:
    # somme des jours de congÃ©s ACCEPTÃ‰S pour cet employÃ©
    total = 0
    for r in REQUESTS:
        if r.get("worker_name", "").strip().lower() != full_name.strip().lower():
            continue
        if r.get("category") == "vacation" and _norm_text(r.get("status")) == "terminee":
            start = _to_date(r.get("start_date"))
            end = _to_date(r.get("end_date"))
            if not start or not end:
                continue
            total += _days_inclusive(start, end)
    # Ã©criture Excel
    set_employee_cumules_in_excel(full_name, total)
    # invalider le cache pour reflÃ©ter la MAJ instantanÃ©ment
    global _employees_mtime
    _employees_mtime = None
    return total

# ================== ENDPOINTS TRAVAILLEUR ==================

@app.post("/worker/login")
def worker_login(data: WorkerLogin):
    employees = get_employees()
    identifier = data.identifier.strip().lower()
    parts = [p for p in identifier.split() if p]
    flipped = " ".join(reversed(parts)) if len(parts) >= 2 else identifier
    def _normalize_code(value: str) -> str:
        raw = str(value or "").strip()
        raw = raw.replace(",", ".")
        if raw.endswith(".0") and raw[:-2].isdigit():
            raw = raw[:-2]
        return raw
    code = _normalize_code(data.code)
    for e in employees:
        full_name = e["full_name"].strip().lower()
        if (full_name == identifier or full_name == flipped) and _normalize_code(e["code"]) == code:
            return {
                "full_name": e["full_name"],
                "email": e["email"],
                "sector": e["sector"],
                "solde": e.get("solde", 0),
            }
    raise HTTPException(status_code=401, detail="Identifiants incorrects")


@app.post("/requests")
def create_request(payload: dict):
    global REQUEST_COUNTER
    req = {
        "id": REQUEST_COUNTER,
        "worker_name": payload.get("worker_name"),
        "worker_email": payload.get("worker_email"),
        "sector": payload.get("sector"),
        "category": payload.get("category"),
        "start_date": payload.get("start_date"),
        "end_date": payload.get("end_date"),
        "reason": payload.get("reason"),
        "title": payload.get("title"),
        "description": payload.get("description"),
        "urgency": payload.get("urgency"),
        "status": "nouvelle",
        "manager_comment": "",
        "created_at": datetime.utcnow(),
        "archived": False,
    }
    REQUESTS.append(req)
    REQUEST_COUNTER += 1
    return req

@app.post("/requests/sick")
async def create_sick_request(
    worker_name: str = Form(...),
    worker_email: str = Form(...),
    sector: str = Form(None),
    start_date: str = Form(...),
    end_date: str = Form(...),
    reason: str = Form(None),
    title: str = Form("CongÃ© maladie"),
    description: str = Form("CongÃ© maladie"),
    urgency: str = Form("normal"),
    attachment: UploadFile | None = File(None),
):
    global REQUEST_COUNTER

    attachment_name = None
    attachment_path = None
    if attachment:
        suffix = Path(attachment.filename or "").suffix
        safe_name = f"{uuid4().hex}{suffix}"
        dest = UPLOAD_DIR / safe_name
        data = await attachment.read()
        dest.write_bytes(data)
        attachment_name = attachment.filename
        attachment_path = str(dest)

    req = {
        "id": REQUEST_COUNTER,
        "worker_name": worker_name,
        "worker_email": worker_email,
        "sector": sector,
        "category": "sick_leave",
        "start_date": start_date,
        "end_date": end_date,
        "reason": reason,
        "title": title,
        "description": description,
        "urgency": urgency,
        "status": "nouvelle",
        "manager_comment": "",
        "created_at": datetime.utcnow(),
        "archived": False,
        "attachment_name": attachment_name,
        "attachment_path": attachment_path,
    }
    REQUESTS.append(req)
    REQUEST_COUNTER += 1
    return req


@app.get("/worker/requests", response_model=List[RequestOut])
def list_worker_requests(full_name: str, email: Optional[str] = None):
    # filtre simple par worker_name (et email si fourni)
    items = [
        r for r in REQUESTS
        if r.get("worker_name", "").strip().lower() == full_name.strip().lower()
    ]
    if email:
        items = [
            r for r in items
            if r.get("worker_email", "").strip().lower() == email.strip().lower()
        ]
    return items

# ================== ENDPOINTS GERANT ==================

@app.post("/admin/login")
def admin_login(payload: dict):
    username = payload.get("username")
    password = payload.get("password")
    if username == "admin" and password == "admin123":  # ou tes vrais identifiants
        return {"token": ADMIN_TOKEN}
    raise HTTPException(status_code=401, detail="Identifiants incorrects")



@app.get("/admin/requests", dependencies=[Depends(require_admin)])
def admin_get_requests(archived: bool = False, worker: str | None = None):
    rs = [r for r in REQUESTS if bool(r.get("archived", False)) == archived]
    if worker:
        rs = [r for r in rs if r["worker_name"].lower() == worker.strip().lower()]
    rs.sort(key=lambda x: (x.get("start_date") or "", x["id"]))
    return rs




@app.patch("/admin/requests/{rid}", dependencies=[Depends(require_admin)])
def admin_update_request(rid: int, payload: dict):
    req = next((r for r in REQUESTS if r["id"] == rid), None)
    if not req:
        raise HTTPException(404, "Demande introuvable")

    desired_status = payload.get("status", req.get("status"))
    desired_comment = payload.get("manager_comment", req.get("manager_comment"))
    norm_status = _norm_text(desired_status)

    warning = None
    conflict_days = []
    over_limit = False
    if norm_status == "terminee":
        sector = req.get("sector") or ""
        req_start = _to_date(req.get("start_date"))
        req_end = _to_date(req.get("end_date"))
        if sector and req_start and req_end:
            counts = _build_sector_day_counts(exclude_req=req)
            for d in _dates_in_range(req_start, req_end):
                if counts.get((sector, d.isoformat()), 0) >= 2:
                    conflict_days.append(d.isoformat())
            if conflict_days:
                if req.get("category") == "vacation":
                    raise HTTPException(
                        status_code=409,
                        detail={
                            "message": (
                                f"Refus: deja 2 personnes du secteur {sector} sont absentes "
                                f"le meme jour ({', '.join(conflict_days)})."
                            ),
                            "conflict_days": conflict_days,
                        },
                    )
                over_limit = True
                warning = (
                    f"Alerte: deja 2 personnes du secteur {sector} sont absentes "
                    f"le meme jour ({', '.join(conflict_days)})."
                )

    req["status"] = desired_status
    req["manager_comment"] = desired_comment
    req["over_limit"] = over_limit if norm_status == "terminee" else False
    if norm_status in ("terminee", "refusee"):
        req["archived"] = True
    if req.get("category") == "vacation" and norm_status == "terminee":
        _recompute_cumules_for(req.get("worker_name", ""))

    return {"ok": True, "warning": warning, "conflict_days": conflict_days}


@app.get("/admin/requests/{rid}/attachment", dependencies=[Depends(require_admin)])
def admin_download_attachment(rid: int):
    req = next((r for r in REQUESTS if r["id"] == rid), None)
    if not req:
        raise HTTPException(404, "Demande introuvable")
    path = req.get("attachment_path")
    name = req.get("attachment_name")
    if not path or not name:
        raise HTTPException(404, "Aucune piÃ¨ce jointe")
    return FileResponse(path, filename=name)



@app.get("/admin/cumules")
def admin_cumules():
    employees = get_employees()  # relit si besoin via cache (inclut secteur + cumules)
    # ne renvoyer que ce qui sert Ã  lâ€™Ã©cran
    return [
        {
            "full_name": e["full_name"],
            "sector": e.get("sector") or "",
            "cumules_excel": int(e.get("cumules", 0) or 0),
        }
        for e in employees
    ]
def get_all_cumules(_=Depends(get_manager_token)):
    # On part du fichier Excel (colonne CongesCumules)
    emps = get_employees()
    # Fournir aussi une "rÃ©fÃ©rence calculÃ©e" pour contrÃ´le visuel
    computed: Dict[str, int] = {}
    for e in emps:
        computed[e["full_name"]] = 0
    for r in REQUESTS:
        if r.get("category") == "vacation" and _norm_text(r.get("status")) == "terminee":
            start = _to_date(r.get("start_date"))
            end = _to_date(r.get("end_date"))
            if not start or not end:
                continue
            name = r.get("worker_name")
            computed[name] = computed.get(name, 0) + _days_inclusive(start, end)

    return [
        {
            "full_name": e["full_name"],
            "email": e["email"],
            "sector": e["sector"],
            "cumules_excel": e["cumules"],   # valeur officielle (Ã©ditable) â€“ ce que voit le RH
            "cumules_calc": computed.get(e["full_name"], 0),  # info pour vÃ©rif
        }
        for e in emps
    ]

@app.patch("/admin/cumules/{full_name}")
def admin_update_cumules(full_name: str, payload: dict):
    days = int(payload.get("days", 0))
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # localiser les colonnes par en-tÃªtes
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    col_name = headers.index("nom")
    col_first = headers.index("prenom")
    col_cum = headers.index("congescumules") if "congescumules" in headers else None
    if col_cum is None:
        raise HTTPException(400, "Colonne 'CongesCumules' absente dans Excel")

    # trouver la ligne employÃ© (match "Prenom Nom")
    target_row = None
    for row in ws.iter_rows(min_row=2):
        nom = str(row[col_name].value or "").strip()
        prenom = str(row[col_first].value or "").strip()
        fn = f"{prenom} {nom}".strip()
        if fn.lower() == full_name.strip().lower():
            target_row = row
            break
    if not target_row:
        raise HTTPException(404, "EmployÃ© introuvable")

    target_row[col_cum].value = days
    wb.save(EXCEL_FILE)
    _upload_excel_to_drive()

    # relire pour renvoyer la valeur officielle
    employees = get_employees()  # relit si besoin via cache (inclut secteur + cumules)
    e = next((x for x in employees if x["full_name"].lower()==full_name.lower()), None)
    return {"full_name": full_name, "cumules_excel": int(e.get("conges_cumules", 0) or 0)}
def set_cumules_for_employee(full_name: str, body: CumuleUpdate, _=Depends(get_manager_token)):
    set_employee_cumules_in_excel(full_name, int(body.days))
    # invalide le cache pour reflÃ©ter
    global _employees_mtime
    _employees_mtime = None
    return {"ok": True, "full_name": full_name, "days": int(body.days)}

@app.get("/admin/employee/{full_name}/requests", dependencies=[Depends(require_admin)])
def admin_employee_history(full_name: str):
    rs = [r for r in REQUESTS if r.get("worker_name","").lower() == full_name.strip().lower()]
    rs.sort(key=lambda x: (x.get("start_date") or "", x.get("id") or 0))
    return rs






